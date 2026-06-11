from fastapi import FastAPI, HTTPException, Query, UploadFile, File
from fastapi.responses import FileResponse, PlainTextResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os
import re
import io
import zipfile
import shutil
from typing import Optional, List

# Excel & PDF libraries
import xlsxwriter
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER

app = FastAPI(
    title="DACT Roster API",
    description="Professional API for employee roster, statistics, rotations with export to Excel/PDF, full CRUD operations and file uploads.",
    version="4.0.0",
    docs_url="/docs",
)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Constants
EXCEL_FILE = "data/OPS Database & Head Count.xlsx"
BACKUP_DIR = "data/backups"
ROTATION_FILES = {
    "STS": "data/STS.xlsx",
    "RTG": "data/RTG.xlsx",
    "TRUCK": "data/TRUCK.xlsx",
    "DECK": "data/DECK.xlsx",
    "RS&EH": "data/RS&EH.xlsx"
}

os.makedirs(BACKUP_DIR, exist_ok=True)
os.makedirs("data", exist_ok=True)
os.makedirs("generated_rotations", exist_ok=True)

# ==========================
# Pydantic Models
# ==========================
class ManpowerRequest(BaseModel):
    required: int
    available: int

class OvertimeRequest(BaseModel):
    required: int
    available: int
    shift_hours: int = 8

class AdvisorRequest(BaseModel):
    operation: str
    units: int
    available: int

class RotationRequest(BaseModel):
    operation: str
    equipment_string: Optional[str] = None
    equipment_numbers: Optional[List[int]] = None

class EmployeeCreate(BaseModel):
    name: str
    phone: str
    hire_date: Optional[str] = None
    position: str
    shift: str

class EmployeeUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    hire_date: Optional[str] = None
    position: Optional[str] = None
    shift: Optional[str] = None

# New models for rotation preview/validation
class RotationPreviewRequest(BaseModel):
    operation: str
    equipment_string: str

class RotationValidateRequest(BaseModel):
    operation: str
    equipment_string: str

# ==========================
# Helper Functions
# ==========================
def parse_equipment_numbers(input_str: str) -> List[int]:
    numbers = re.findall(r'\d+', input_str)
    return [int(num) for num in numbers]

def get_rotation_workbook(operation: str):
    """Load workbook for given operation, return (wb, operation_upper)."""
    op_upper = operation.upper()
    if op_upper not in ROTATION_FILES:
        raise HTTPException(status_code=400, detail="Invalid operation")
    path = ROTATION_FILES[op_upper]
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"Rotation file not found for {op_upper}")
    wb = load_workbook(path)
    return wb, op_upper

def load_employee_data() -> List[dict]:
    """Load employee data from first sheet of main Excel file."""
    if not os.path.exists(EXCEL_FILE):
        # Return empty list if file doesn't exist (will be created on first save)
        return []
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    # Find first row where first column is a number (ID)
    start_row = 0
    for i, row in enumerate(rows):
        if row[0] is not None and str(row[0]).strip().isdigit():
            start_row = i
            break
    employees = []
    for row in rows[start_row:]:
        if len(row) < 6:
            continue
        id_val = row[0]
        if id_val is None or not str(id_val).strip().isdigit():
            continue
        employees.append({
            "ID": str(id_val).strip(),
            "Name": str(row[1]).strip() if row[1] else "",
            "Phone": str(row[2]).strip() if row[2] else "",
            "HireDate": str(row[3]) if row[3] else "",
            "Position": str(row[4]).strip() if row[4] else "",
            "Shift": str(row[5]).strip().upper() if row[5] else ""
        })
    return employees

def save_employee_data(employees: List[dict]):
    """Save employee data back to first sheet of main Excel file, preserving other sheets."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"
    else:
        wb = load_workbook(EXCEL_FILE)
        # Ensure there is at least one sheet
        if len(wb.worksheets) == 0:
            wb.create_sheet("Employees")
        ws = wb.worksheets[0]
        # Delete all rows after header (1) but keep the sheet
        ws.delete_rows(2, ws.max_row)
    
    # Write headers
    headers = ["ID", "Name", "Phone", "HireDate", "Position", "Shift"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    
    # Write data
    for row_idx, emp in enumerate(employees, start=2):
        ws.cell(row=row_idx, column=1, value=emp["ID"])
        ws.cell(row=row_idx, column=2, value=emp["Name"])
        ws.cell(row=row_idx, column=3, value=emp["Phone"])
        ws.cell(row=row_idx, column=4, value=emp["HireDate"])
        ws.cell(row=row_idx, column=5, value=emp["Position"])
        ws.cell(row=row_idx, column=6, value=emp["Shift"])
    
    wb.save(EXCEL_FILE)

def get_next_id(employees: List[dict]) -> int:
    if not employees:
        return 1000
    ids = [int(emp["ID"]) for emp in employees if emp["ID"].isdigit()]
    return max(ids) + 1

def backup_file(file_path: str) -> Optional[str]:
    """Create a timestamped backup of a file. Returns backup path or None."""
    if os.path.exists(file_path):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = os.path.basename(file_path).replace(".xlsx", f"_{timestamp}.xlsx")
        backup_path = os.path.join(BACKUP_DIR, backup_name)
        shutil.copy2(file_path, backup_path)
        return backup_path
    return None

def get_statistics():
    data = load_employee_data()
    shift_counts = {"A":0,"B":0,"C":0,"D":0,"DW":0}
    pos_counts = {}
    for emp in data:
        sh = emp["Shift"]
        if sh in shift_counts:
            shift_counts[sh] += 1
        pos = emp["Position"]
        if pos:
            pos_counts[pos] = pos_counts.get(pos, 0) + 1
    return {
        "total": len(data),
        "shifts": shift_counts,
        "positions": pos_counts
    }

def get_positions_by_shift():
    data = load_employee_data()
    shifts = ["A","B","C","D","DW"]
    result = {}
    for emp in data:
        pos = emp["Position"]
        shift = emp["Shift"]
        if not pos or shift not in shifts:
            continue
        if pos not in result:
            result[pos] = {s:0 for s in shifts}
        result[pos][shift] += 1
    return result

# ------------------------------
# Export helpers (Excel & PDF)
# ------------------------------
def export_employees_excel():
    data = load_employee_data()
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Employees")
    header_format = workbook.add_format({'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white', 'border': 1})
    cell_format = workbook.add_format({'border': 1})
    headers = ["ID", "Name", "Phone", "Hire Date", "Position", "Shift"]
    for col, h in enumerate(headers):
        worksheet.write(0, col, h, header_format)
    for row, emp in enumerate(data, start=1):
        worksheet.write(row, 0, emp["ID"], cell_format)
        worksheet.write(row, 1, emp["Name"], cell_format)
        worksheet.write(row, 2, emp["Phone"], cell_format)
        worksheet.write(row, 3, emp["HireDate"], cell_format)
        worksheet.write(row, 4, emp["Position"], cell_format)
        worksheet.write(row, 5, emp["Shift"], cell_format)
    worksheet.set_column(0, 0, 12)
    worksheet.set_column(1, 1, 35)
    worksheet.set_column(2, 2, 15)
    worksheet.set_column(3, 3, 12)
    worksheet.set_column(4, 4, 30)
    worksheet.set_column(5, 5, 8)
    workbook.close()
    output.seek(0)
    return output

def export_stats_excel():
    stats = get_statistics()
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    # Shift sheet
    ws1 = workbook.add_worksheet("Shift Statistics")
    header_fmt = workbook.add_format({'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white'})
    ws1.write(0, 0, "Shift", header_fmt)
    ws1.write(0, 1, "Count", header_fmt)
    row = 1
    for sh, cnt in stats["shifts"].items():
        ws1.write(row, 0, sh)
        ws1.write(row, 1, cnt)
        row += 1
    ws1.write(row, 0, "Total")
    ws1.write(row, 1, stats["total"])
    # Position sheet
    ws2 = workbook.add_worksheet("Position Statistics")
    ws2.write(0, 0, "Position", header_fmt)
    ws2.write(0, 1, "Count", header_fmt)
    sorted_pos = sorted(stats["positions"].items(), key=lambda x: x[1], reverse=True)
    for r, (pos, cnt) in enumerate(sorted_pos, start=1):
        ws2.write(r, 0, pos)
        ws2.write(r, 1, cnt)
    workbook.close()
    output.seek(0)
    return output

def export_positions_by_shift_excel():
    data = get_positions_by_shift()
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Positions by Shift")
    header_fmt = workbook.add_format({'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white', 'border': 1})
    cell_fmt = workbook.add_format({'border': 1})
    shifts = ["A","B","C","D","DW"]
    worksheet.write(0, 0, "Position", header_fmt)
    for col, sh in enumerate(shifts, start=1):
        worksheet.write(0, col, f"Shift {sh}", header_fmt)
    worksheet.write(0, len(shifts)+1, "Total", header_fmt)
    row = 1
    for pos, sh_counts in data.items():
        worksheet.write(row, 0, pos, cell_fmt)
        total = 0
        for col, sh in enumerate(shifts, start=1):
            cnt = sh_counts.get(sh, 0)
            worksheet.write(row, col, cnt, cell_fmt)
            total += cnt
        worksheet.write(row, len(shifts)+1, total, cell_fmt)
        row += 1
    worksheet.set_column(0, 0, 30)
    for i in range(1, len(shifts)+2):
        worksheet.set_column(i, i, 10)
    workbook.close()
    output.seek(0)
    return output

def export_rotation_excel(operation: str, units: int, rotation_data: List[dict]):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet(f"{operation}_{units}units")
    header_fmt = workbook.add_format({'bold': True, 'bg_color': '#4F81BD', 'font_color': 'white', 'border': 1})
    cell_fmt = workbook.add_format({'border': 1})
    if rotation_data:
        headers = list(rotation_data[0].keys())
        for col, h in enumerate(headers):
            worksheet.write(0, col, h, header_fmt)
        for row, record in enumerate(rotation_data, start=1):
            for col, key in enumerate(headers):
                worksheet.write(row, col, record.get(key, ""), cell_fmt)
    workbook.close()
    output.seek(0)
    return output

def create_pdf(title, table_data, headers, col_widths=None):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16, spaceAfter=20)
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 12))
    tbl_data = [headers] + table_data
    if col_widths is None:
        col_widths = [inch] * len(headers)
    table = Table(tbl_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4F81BD')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer

def export_employees_pdf():
    data = load_employee_data()
    headers = ["ID", "Name", "Phone", "Hire Date", "Position", "Shift"]
    table_data = [[emp["ID"], emp["Name"], emp["Phone"], emp["HireDate"], emp["Position"], emp["Shift"]] for emp in data]
    col_widths = [0.8*inch, 2.2*inch, 1.2*inch, 1.0*inch, 1.5*inch, 0.6*inch]
    return create_pdf("Employee Roster Report", table_data, headers, col_widths)

def export_stats_pdf():
    stats = get_statistics()
    shift_headers = ["Shift", "Count"]
    shift_data = [[sh, cnt] for sh, cnt in stats["shifts"].items()]
    shift_data.append(["Total", stats["total"]])
    pos_headers = ["Position", "Count"]
    pos_data = [[pos, cnt] for pos, cnt in sorted(stats["positions"].items(), key=lambda x: x[1], reverse=True)]
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=16, spaceAfter=20)
    elements.append(Paragraph("Statistics Report", title_style))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Shift Statistics", styles['Heading2']))
    shift_table = Table([shift_headers] + shift_data, colWidths=[1.5*inch, 1*inch])
    shift_table.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4F81BD')), ('TEXTCOLOR', (0,0), (-1,0), colors.white)]))
    elements.append(shift_table)
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Position Statistics", styles['Heading2']))
    pos_table = Table([pos_headers] + pos_data, colWidths=[3*inch, 1*inch])
    pos_table.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4F81BD')), ('TEXTCOLOR', (0,0), (-1,0), colors.white)]))
    elements.append(pos_table)
    doc.build(elements)
    buffer.seek(0)
    return buffer

def export_positions_by_shift_pdf():
    data = get_positions_by_shift()
    shifts = ["A","B","C","D","DW"]
    headers = ["Position"] + [f"Shift {s}" for s in shifts] + ["Total"]
    table_data = []
    for pos, sh_counts in data.items():
        row = [pos] + [sh_counts.get(s, 0) for s in shifts] + [sum(sh_counts.values())]
        table_data.append(row)
    table_data.sort(key=lambda x: x[-1], reverse=True)
    col_widths = [2.0*inch] + [0.6*inch]*len(shifts) + [0.6*inch]
    return create_pdf("Positions Distribution by Shift", table_data, headers, col_widths)

def export_rotation_pdf(operation: str, units: int, rotation_data: List[dict]):
    if not rotation_data:
        raise ValueError("No rotation data")
    headers = list(rotation_data[0].keys())
    table_data = [[str(record.get(h, "")) for h in headers] for record in rotation_data]
    col_widths = [0.8*inch] * len(headers)
    title = f"Rotation Pattern - {operation} ({units} units)"
    return create_pdf(title, table_data, headers, col_widths)

# ==========================
# Endpoints
# ==========================
@app.get("/")
def home():
    return {"status": "running", "system": "DACT Roster API", "docs": "/docs"}

# --------------------------
# Basic employee data
# --------------------------
@app.get("/employees")
def employees():
    try:
        data = load_employee_data()
        return {"employees_count": len(data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/employee/{emp_id}")
def get_employee(emp_id: str):
    try:
        data = load_employee_data()
        for emp in data:
            if emp["ID"] == emp_id:
                return emp
        raise HTTPException(status_code=404, detail="Employee not found")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/employee/search/{name}")
def search_employee(name: str):
    try:
        data = load_employee_data()
        results = []
        for emp in data:
            if name.lower() in emp["Name"].lower():
                results.append({"id": emp["ID"], "name": emp["Name"], "position": emp["Position"], "shift": emp["Shift"]})
        return results
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/shift/{shift}")
def search_shift(shift: str):
    try:
        data = load_employee_data()
        shift_upper = shift.upper()
        results = []
        for emp in data:
            if emp["Shift"] == shift_upper:
                results.append({"id": emp["ID"], "name": emp["Name"], "position": emp["Position"], "shift": emp["Shift"]})
        return results
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/position/{position}")
def search_position(position: str):
    try:
        data = load_employee_data()
        pos_lower = position.lower()
        results = []
        for emp in data:
            if pos_lower in emp["Position"].lower():
                results.append({"id": emp["ID"], "name": emp["Name"], "position": emp["Position"], "shift": emp["Shift"]})
        return results
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --------------------------
# Statistics
# --------------------------
@app.get("/stats")
def stats():
    try:
        s = get_statistics()
        return {
            "total_employees": s["total"],
            "shift_a": s["shifts"]["A"],
            "shift_b": s["shifts"]["B"],
            "shift_c": s["shifts"]["C"],
            "shift_d": s["shifts"]["D"],
            "shift_dw": s["shifts"]["DW"]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/positions/stats")
def positions_stats():
    try:
        s = get_statistics()
        sorted_pos = dict(sorted(s["positions"].items(), key=lambda x: x[1], reverse=True))
        return {"total_positions": len(sorted_pos), "positions": sorted_pos}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/positions/by_shift")
def positions_by_shift():
    try:
        data = get_positions_by_shift()
        return {"shifts": ["A","B","C","D","DW"], "data": data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/shifts")
def list_shifts():
    return {"shifts": ["A","B","C","D","DW"]}

@app.get("/positions")
def list_positions():
    try:
        data = load_employee_data()
        positions = sorted(set(emp["Position"] for emp in data if emp["Position"]))
        return {"positions": positions}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/operations")
def operations():
    return {"operations": list(ROTATION_FILES.keys())}

# --------------------------
# Export endpoints (Excel & PDF)
# --------------------------
@app.get("/export/employees")
def export_employees_endpoint(format: str = Query(..., pattern="^(excel|pdf)$")):
    try:
        if format == "excel":
            file = export_employees_excel()
            return StreamingResponse(file, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=employees_{datetime.now().strftime('%Y%m%d')}.xlsx"})
        else:
            file = export_employees_pdf()
            return StreamingResponse(file, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=employees_{datetime.now().strftime('%Y%m%d')}.pdf"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/export/stats")
def export_stats_endpoint(format: str = Query(..., pattern="^(excel|pdf)$")):
    try:
        if format == "excel":
            file = export_stats_excel()
            return StreamingResponse(file, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=stats_{datetime.now().strftime('%Y%m%d')}.xlsx"})
        else:
            file = export_stats_pdf()
            return StreamingResponse(file, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=stats_{datetime.now().strftime('%Y%m%d')}.pdf"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/export/positions_by_shift")
def export_positions_by_shift_endpoint(format: str = Query(..., pattern="^(excel|pdf)$")):
    try:
        if format == "excel":
            file = export_positions_by_shift_excel()
            return StreamingResponse(file, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=positions_by_shift_{datetime.now().strftime('%Y%m%d')}.xlsx"})
        else:
            file = export_positions_by_shift_pdf()
            return StreamingResponse(file, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=positions_by_shift_{datetime.now().strftime('%Y%m%d')}.pdf"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/export/rotation/{operation}/{units}")
def export_rotation_endpoint(operation: str, units: int, format: str = Query(..., pattern="^(excel|pdf)$")):
    operation = operation.upper()
    if operation not in ROTATION_FILES:
        raise HTTPException(status_code=400, detail="Invalid operation")
    try:
        import pandas as pd
        df = pd.read_excel(ROTATION_FILES[operation], sheet_name=str(units), header=None)
        rotation = df.fillna("").to_dict(orient="records")
        if format == "excel":
            file = export_rotation_excel(operation, units, rotation)
            return StreamingResponse(file, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=rotation_{operation}_{units}_{datetime.now().strftime('%Y%m%d')}.xlsx"})
        else:
            file = export_rotation_pdf(operation, units, rotation)
            return StreamingResponse(file, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=rotation_{operation}_{units}_{datetime.now().strftime('%Y%m%d')}.pdf"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/export/all")
def export_all():
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        zipf.writestr("employees.xlsx", export_employees_excel().getvalue())
        zipf.writestr("stats.xlsx", export_stats_excel().getvalue())
        zipf.writestr("positions_by_shift.xlsx", export_positions_by_shift_excel().getvalue())
        # Add rotation files if they exist
        for op, path in ROTATION_FILES.items():
            if os.path.exists(path):
                try:
                    import pandas as pd
                    for sheet in range(1, 20):
                        try:
                            df = pd.read_excel(path, sheet_name=str(sheet), header=None)
                            rot = df.fillna("").to_dict(orient="records")
                            file = export_rotation_excel(op, sheet, rot)
                            zipf.writestr(f"rotation_{op}_{sheet}.xlsx", file.getvalue())
                        except:
                            break
                except:
                    pass
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/zip", headers={"Content-Disposition": f"attachment; filename=all_reports_{datetime.now().strftime('%Y%m%d')}.zip"})

# --------------------------
# CRUD for Employees
# --------------------------
@app.post("/employees", status_code=201)
def create_employee(emp: EmployeeCreate):
    try:
        employees = load_employee_data()
        new_id = str(get_next_id(employees))
        new_emp = {
            "ID": new_id,
            "Name": emp.name,
            "Phone": emp.phone,
            "HireDate": emp.hire_date or "",
            "Position": emp.position,
            "Shift": emp.shift.upper()
        }
        employees.append(new_emp)
        backup_file(EXCEL_FILE)
        save_employee_data(employees)
        return {"message": "Employee added", "employee": new_emp}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/employee/{emp_id}")
def update_employee(emp_id: str, emp_update: EmployeeUpdate):
    try:
        employees = load_employee_data()
        found = False
        for emp in employees:
            if emp["ID"] == emp_id:
                if emp_update.name is not None:
                    emp["Name"] = emp_update.name
                if emp_update.phone is not None:
                    emp["Phone"] = emp_update.phone
                if emp_update.hire_date is not None:
                    emp["HireDate"] = emp_update.hire_date
                if emp_update.position is not None:
                    emp["Position"] = emp_update.position
                if emp_update.shift is not None:
                    emp["Shift"] = emp_update.shift.upper()
                found = True
                break
        if not found:
            raise HTTPException(status_code=404, detail="Employee not found")
        backup_file(EXCEL_FILE)
        save_employee_data(employees)
        return {"message": "Employee updated"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/employee/{emp_id}")
def delete_employee(emp_id: str):
    try:
        employees = load_employee_data()
        new_employees = [emp for emp in employees if emp["ID"] != emp_id]
        if len(new_employees) == len(employees):
            raise HTTPException(status_code=404, detail="Employee not found")
        backup_file(EXCEL_FILE)
        save_employee_data(new_employees)
        return {"message": "Employee deleted"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --------------------------
# Upload rotation file for a specific operation
# --------------------------
@app.post("/rotation/upload/{operation}")
async def upload_rotation(operation: str, file: UploadFile = File(...)):
    operation = operation.upper()
    if operation not in ROTATION_FILES:
        raise HTTPException(status_code=400, detail="Invalid operation")
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files allowed")
    try:
        target_path = ROTATION_FILES[operation]
        backup_file(target_path)
        content = await file.read()
        with open(target_path, "wb") as buffer:
            buffer.write(content)
        return {"message": f"Rotation file for {operation} uploaded successfully", "path": target_path}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --------------------------
# Upload full database (replace main Excel file)
# --------------------------
@app.post("/database/upload")
async def upload_database(file: UploadFile = File(...)):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files allowed")
    try:
        backup_file(EXCEL_FILE)
        content = await file.read()
        with open(EXCEL_FILE, "wb") as buffer:
            buffer.write(content)
        return {"message": "Database replaced successfully", "path": EXCEL_FILE}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --------------------------
# Planning and Rotation endpoints (existing)
# --------------------------
@app.post("/manpower-gap")
def manpower_gap(data: ManpowerRequest):
    gap = data.required - data.available
    return {"required": data.required, "available": data.available, "gap": gap, "status": "Shortage" if gap > 0 else "Covered"}

@app.post("/coverage")
def coverage(data: ManpowerRequest):
    if data.required == 0:
        return {"coverage_percent": 0}
    return {"coverage_percent": round((data.available / data.required) * 100, 2)}

@app.post("/overtime")
def overtime(data: OvertimeRequest):
    missing = max(data.required - data.available, 0)
    return {"missing_workers": missing, "overtime_hours": missing * data.shift_hours}

@app.get("/rotation/{operation}/{units}")
def get_rotation(operation: str, units: int):
    operation = operation.upper()
    if operation not in ROTATION_FILES:
        raise HTTPException(status_code=400, detail="Invalid operation")
    try:
        import pandas as pd
        df = pd.read_excel(ROTATION_FILES[operation], sheet_name=str(units), header=None)
        return {"operation": operation, "units": units, "rows": len(df), "rotation": df.fillna("").to_dict(orient="records")}
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Rotation file for {operation} not found")
    except ValueError:
        raise HTTPException(status_code=404, detail=f"No sheet found for {units} units")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/rotation/custom")
def custom_rotation(data: RotationRequest):
    operation = data.operation.upper()
    if operation not in ROTATION_FILES:
        raise HTTPException(status_code=400, detail="Invalid operation")
    equipment_numbers = []
    if data.equipment_numbers is not None:
        equipment_numbers = data.equipment_numbers.copy()
    elif data.equipment_string is not None:
        equipment_numbers = parse_equipment_numbers(data.equipment_string)
    else:
        raise HTTPException(status_code=400, detail="Provide 'equipment_numbers' or 'equipment_string'")
    if not equipment_numbers:
        raise HTTPException(status_code=400, detail="No valid equipment numbers provided")
    equipment_numbers.sort()
    units = len(equipment_numbers)
    try:
        wb = load_workbook(ROTATION_FILES[operation])
        sheet_name = str(units)
        if sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=404, detail=f"No sheet found for {units} units")
        ws = wb[sheet_name]
        mapping = {str(i): str(eq) for i, eq in enumerate(equipment_numbers, start=1)}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value).strip()
                if val in mapping:
                    cell.value = mapping[val]
        for sheet in list(wb.sheetnames):
            if sheet != sheet_name:
                wb.remove(wb[sheet])
        os.makedirs("generated_rotations", exist_ok=True)
        filename = f"{operation}_{units}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join("generated_rotations", filename)
        wb.save(output_path)
        wb.close()
        return FileResponse(output_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=filename)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"Rotation file for {operation} not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==========================
# Rotation Preview & Validation Endpoints (NEW)
# ==========================

@app.post("/rotation/preview")
def rotation_preview(request: RotationPreviewRequest):
    """
    Preview how equipment numbers will be processed for rotation.
    Returns sorted numbers, duplicates, units, and sheet availability.
    """
    operation = request.operation.upper()
    if operation not in ROTATION_FILES:
        raise HTTPException(status_code=400, detail="Invalid operation")
    
    # Parse equipment numbers from string (handles any delimiters)
    numbers = parse_equipment_numbers(request.equipment_string)
    if not numbers:
        raise HTTPException(status_code=400, detail="No valid equipment numbers provided")
    
    # Sort ascending
    sorted_numbers = sorted(numbers)
    
    # Detect duplicates (numbers that appear more than once)
    seen = {}
    duplicates = []
    for n in numbers:
        if n in seen:
            if n not in duplicates:
                duplicates.append(n)
        else:
            seen[n] = 1
    
    units = len(sorted_numbers)
    
    # Check if the rotation sheet exists
    sheet_exists = False
    try:
        wb = load_workbook(ROTATION_FILES[operation])
        sheet_exists = str(units) in wb.sheetnames
        wb.close()
    except:
        sheet_exists = False
    
    return {
        "operation": operation,
        "units": units,
        "sorted_equipment": sorted_numbers,
        "duplicates": duplicates,
        "sheet_exists": sheet_exists,
        "sheet": str(units) if sheet_exists else None
    }


@app.post("/rotation/validate")
def rotation_validate(request: RotationValidateRequest):
    """
    Validate if a rotation request can be processed.
    Checks: operation exists, sheet exists, no duplicate numbers.
    """
    operation = request.operation.upper()
    errors = []
    operation_valid = operation in ROTATION_FILES
    if not operation_valid:
        errors.append(f"Invalid operation '{request.operation}'. Allowed: {list(ROTATION_FILES.keys())}")
    
    # Parse numbers
    numbers = parse_equipment_numbers(request.equipment_string)
    if not numbers:
        errors.append("No valid equipment numbers provided")
    
    # Check duplicates
    seen = set()
    duplicates = []
    for n in numbers:
        if n in seen:
            if n not in duplicates:
                duplicates.append(n)
        else:
            seen.add(n)
    duplicate_numbers = duplicates
    
    sheet_valid = False
    if operation_valid and numbers:
        try:
            wb = load_workbook(ROTATION_FILES[operation])
            units = len(numbers)
            sheet_valid = str(units) in wb.sheetnames
            if not sheet_valid:
                errors.append(f"No rotation sheet found for {units} units")
            wb.close()
        except Exception as e:
            errors.append(f"Cannot read rotation file: {str(e)}")
    else:
        if operation_valid:
            errors.append("Cannot validate sheet due to missing numbers")
    
    if duplicate_numbers:
        errors.append(f"Duplicate equipment numbers found: {duplicate_numbers}")
    
    valid = (operation_valid and sheet_valid and not duplicate_numbers and numbers)
    
    return {
        "valid": valid,
        "operation_valid": operation_valid,
        "sheet_valid": sheet_valid,
        "duplicate_numbers": duplicate_numbers,
        "errors": errors
    }


@app.get("/rotation/sheets/{operation}")
def rotation_available_sheets(operation: str):
    """
    Return all available sheet names (unit counts) for a given operation.
    """
    operation = operation.upper()
    if operation not in ROTATION_FILES:
        raise HTTPException(status_code=400, detail="Invalid operation")
    
    file_path = ROTATION_FILES[operation]
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail=f"Rotation file for {operation} not found")
    
    try:
        wb = load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return {
            "operation": operation,
            "available_sheets": sheets
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading workbook: {str(e)}")


@app.post("/advisor")
def advisor(data: AdvisorRequest):
    required = data.units * 2
    gap = max(required - data.available, 0)
    return {"operation": data.operation, "units": data.units, "required": required, "available": data.available, "gap": gap, "recommendation": "Coverage OK" if gap == 0 else f"Need {gap} more employees"}